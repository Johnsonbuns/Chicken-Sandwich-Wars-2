#!/usr/bin/env python3
"""
Packs research/sheet/*.csv into one .xlsx that Google Sheets imports cleanly.

    pip install openpyxl && python3 scripts/research-sheet-xlsx.py

The CSVs are the dependency-free artefact and the source of truth for structure; this
adds only what a spreadsheet needs to be usable by a human — frozen headers, a purpose
line on every tab, zone colouring, and dropdowns on the columns where a free-text value
would quietly break the reconciliation back into Supabase.
"""
import csv, json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SHEET = os.path.join(ROOT, 'research', 'sheet')
OUT = os.path.join(ROOT, 'research', 'CSW-Master-Research-Database.xlsx')

ZONE = {
    'A · IDENTITY': ('1F3A5F', 'DCE6F1', '1F4E79'),   # tab colour, header fill, header text
    'B · OBSERVED': ('1E5631', 'DDEBE0', '1E5631'),
    'C · DERIVED':  ('7F4F00', 'FBE9D0', '7F4F00'),
}

# Vocabularies are production's, not new ones. A sheet that invents its own words for
# "verified" turns reconciliation into a translation project.
VOCAB = {
    'doc_type': 'sec_filing,company_release,trade_press,data_provider,brokerage_report,fdd,court_filing,government,other',
    'kind': 'franchisor,operator,private_equity,reit,public_holding,broker,lender,developer,supplier,other',
    'status_company': 'active,acquiring,divesting,refranchising,newly_formed,stable,acquired,restructuring,liquidated',
    'verification': 'unverified,desk_reviewed,verified,rejected',
    'extraction_confidence': 'confirmed,high,medium,low,unknown',
    'match_confidence': 'confirmed,high,medium,low,unknown',
    'confidence': 'confirmed,high,medium,low,unknown',
    'derivation': 'reported,derived,franchisee_reported,company_guidance',
    'visibility': 'public,internal,confidential',
    'roster_type': 'current,departed',
    'priority': 'P0,P1,P2,P3',
    'yesno': 'yes,no,partial',
    'severity': 'high,medium,low',
    'event': 'new,exited,grew,shrank,flat,renamed,merged',
    'signal_type': ('unit_growth,unit_decline,system_contraction,portfolio_transfer,acquisition,'
                    'divestiture,bankruptcy,litigation,sale_leaseback,refranchising,market_entry,'
                    'market_exit,executive_change,development_agreement,distress_composite'),
    'direction': 'positive,negative,neutral',
    'basis': 'all_units,franchised_only,company_only,top_quartile,top_half,top_decile,mature_units,other',
    'match_method': 'canonical,exact,normalized,fuzzy,manual,unresolved',
}

# column name -> vocabulary key
COLVOCAB = {
    'doc_type': 'doc_type', 'kind': 'kind', 'verification': 'verification',
    'extraction_confidence': 'extraction_confidence', 'match_confidence': 'match_confidence',
    'confidence': 'confidence', 'derivation': 'derivation', 'visibility': 'visibility',
    'roster_type': 'roster_type', 'priority': 'priority', 'severity': 'severity',
    'event': 'event', 'signal_type': 'signal_type', 'direction': 'direction',
    'basis': 'basis', 'match_method': 'match_method',
    'file_held': 'yesno', 'item19_extracted': 'yesno', 'item20_extracted': 'yesno',
    'roster_extracted': 'yesno', 'in_production_csw': 'yesno', 'is_chicken': 'yesno',
    'ready_to_submit': 'yesno', 'is_second_generation': 'yesno',
}

WIDE = {'notes', 'purpose', 'raw_quote', 'description', 'why_valuable', 'rationale',
        'title', 'analysis', 'resolution', 'payload_json', 'sources_json', 'task',
        'address_raw', 'address_normalized', 'url', 'brands_operated', 'geography'}

manifest = json.load(open(os.path.join(SHEET, '_manifest.json')))
wb = Workbook()
wb.remove(wb.active)

thin = Side(style='thin', color='C9CDD3')

# ---------------------------------------------------------------- README ----
ws = wb.create_sheet('README')
ws.sheet_properties.tabColor = '000000'
readme = [
    ('CSW MASTER RESEARCH DATABASE', 'h1'),
    ('The staging and intelligence layer behind ChickenSandwichWars.com. Research accumulates here first; '
     'nothing here is on the site.', 'p'),
    ('', 'p'),
    ('THE ONE RULE', 'h2'),
    ('Every figure carries a publisher, a URL and an as-of date. Where a number has not been published, '
     'the site shows "—" rather than an estimate. That rule is the product, and it applies here too: an '
     'unsourced value in this workbook is a liability, not a placeholder.', 'p'),
    ('', 'p'),
    ('THREE ZONES, AND THE DISCIPLINE IS THE POINT', 'h2'),
    ('A · IDENTITY (blue) — the spine. Persistent ids that everything else references. '
     'Never renumber an id; a changed name gets a new row in operator_aliases, not an edit.', 'p'),
    ('B · OBSERVED (green) — append-only. What a document actually said, with its citation. '
     'A correction is a NEW row pointing at the old one, never an overwrite. This mirrors how '
     'production `facts` supersedes rather than updates.', 'p'),
    ('C · DERIVED (amber) — computed from A + B. Do not hand-edit; regenerating overwrites it.', 'p'),
    ('', 'p'),
    ('People and agents write to A and B. C is output. That is what stops the usual spreadsheet '
     'death spiral where nobody can tell which cell is the truth.', 'p'),
    ('', 'p'),
    ('WHY FDD DATA IS A ROSTER, NOT AN EVENT LOG', 'h2'),
    ('An FDD franchisee list is exhaustive as of a date. That exhaustiveness is the whole asset: it is '
     'what makes a DISAPPEARANCE detectable. An event log can never answer "which operators exited", '
     'because the absence of an event is not an event. So we store rosters (fdd_roster) and COMPUTE the '
     'events (roster_diff, signals). Roster in, events out.', 'p'),
    ('', 'p'),
    ('The highest-value page in any FDD is the departed-franchisee list — franchisees whose outlets were '
     'terminated, cancelled, not renewed or which ceased operations in the last fiscal year, with contact '
     'details. The FTC Franchise Rule (16 CFR 436.5(t)) requires it annually, per brand. It is a distress '
     'and lead list handed over by federal law. Capture it as roster_type = departed.', 'p'),
    ('', 'p'),
    ('HOW A ROW REACHES THE WEBSITE', 'h2'),
    ('observations  →  publish_candidates  →  POST /api/agent  →  desk review queue  →  approve  →  '
     'Publish to site  →  data/*.json  →  Vercel', 'mono'),
    ('', 'p'),
    ('Two things people get wrong. Approving is NOT publishing: the build reads data/*.json and never '
     'touches the database, so a figure reaches the site only when someone presses "Publish to site" in '
     'the desk. And an agent cannot write canonical data at all — its key reaches four review RPCs and '
     'nothing else. Both are enforced in Postgres, not by convention.', 'p'),
    ('', 'p'),
    ('CONFLICTS ARE NEVER RESOLVED BY DELETING A ROW', 'h2'),
    ('Two FDD aggregators put Zaxby\'s AUV at $2,847,345 and $2,544,354 — a $303k spread on a figure that '
     'decides a rank. Reachable is not sourced. Disagreements live in `conflicts` until a primary document '
     'settles them.', 'p'),
    ('', 'p'),
    ('REGENERATING', 'h2'),
    ('node scripts/research-sheet.js        # re-seed the CSVs from live CSW data', 'mono'),
    ('python3 scripts/research-sheet-xlsx.py # repack this workbook', 'mono'),
]
r = 1
for text, kind in readme:
    c = ws.cell(row=r, column=1, value=text)
    if kind == 'h1':
        c.font = Font(size=18, bold=True, color='1F3A5F')
    elif kind == 'h2':
        c.font = Font(size=12, bold=True, color='1F3A5F')
    elif kind == 'mono':
        c.font = Font(size=10, name='Courier New', color='444444')
    else:
        c.font = Font(size=11, color='222222')
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = None if kind in ('h1', 'h2') else (15 if not text else max(15, 15 * (len(text) // 105 + 1)))
    r += 1
ws.column_dimensions['A'].width = 115

# ------------------------------------------------------------- data tabs ----
for spec in manifest:
    name, zone, purpose, cols = spec['name'], spec['zone'], spec['purpose'], spec['columns']
    tabc, fill, txt = ZONE[zone]
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tabc

    ws.cell(row=1, column=1, value=zone.split(' · ')[1] + ' — ' + purpose)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(cols), 2))
    ws.cell(row=1, column=1).font = Font(size=10, italic=True, color=txt)
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[1].height = 42

    for i, col in enumerate(cols, start=1):
        c = ws.cell(row=2, column=i, value=col)
        c.font = Font(bold=True, size=10, color=txt)
        c.fill = PatternFill('solid', fgColor=fill)
        c.alignment = Alignment(wrap_text=True, vertical='bottom')
        c.border = Border(bottom=Side(style='medium', color=tabc))
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 46 if col in WIDE else (
            13 if len(col) < 12 else min(26, len(col) + 4))

    with open(os.path.join(SHEET, name + '.csv'), newline='', encoding='utf-8') as fh:
        rows = list(csv.reader(fh))[1:]
    for ri, row in enumerate(rows, start=3):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(wrap_text=cols[ci - 1] in WIDE, vertical='top')
            c.font = Font(size=10)
            c.border = Border(bottom=thin)

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(cols))}{max(len(rows) + 2, 3)}'

    for i, col in enumerate(cols, start=1):
        key = COLVOCAB.get(col)
        if not key:
            continue
        dv = DataValidation(type='list', formula1='"' + VOCAB[key] + '"',
                            allow_blank=True, showDropDown=False)
        dv.error = 'Use one of the listed values — these match the production database enums.'
        dv.errorTitle = 'Controlled vocabulary'
        ws.add_data_validation(dv)
        letter = get_column_letter(i)
        dv.add(f'{letter}3:{letter}2000')

wb.save(OUT)
print('wrote', os.path.relpath(OUT, ROOT), '·', len(wb.sheetnames), 'tabs')
