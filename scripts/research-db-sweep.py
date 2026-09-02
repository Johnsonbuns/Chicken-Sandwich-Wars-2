#!/usr/bin/env python3
"""
Reads the research workbook, works out what is new, and prepares it for CSW.

    python3 scripts/research-db-sweep.py research/CSW-Research-Database.xlsx
    python3 scripts/research-db-sweep.py <file> --write   # also fill the » columns

This never modifies your workbook in place. With --write it saves a copy alongside it
(`...-swept.xlsx`) with only the » machine columns filled in; every column you own comes
through untouched. Diff the two, keep whichever you prefer.

What a sweep does:

  1. assigns ids to rows that do not have one
  2. matches operator and brand names against what we already know, including aliases,
     and reports what it could not match rather than guessing
  3. flags rows whose value contradicts something CSW already publishes
  4. shapes the rows that are ready into a POST /api/agent findings file

Nothing here writes to the website. The findings file goes to the review queue, a human
approves it in the desk, and it reaches the site only when someone presses Publish.
"""
import json, os, re, sys, datetime, difflib
from openpyxl import load_workbook

TODAY = datetime.date.today().isoformat()
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

if len(sys.argv) < 2:
    sys.exit(__doc__)
PATH = sys.argv[1]
WRITE = '--write' in sys.argv

wb = load_workbook(PATH)

def table(name):
    """Rows of a tab as dicts, with the sheet row number attached."""
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    cols = [c.value for c in ws[2]]
    out = []
    for r in ws.iter_rows(min_row=3, values_only=False):
        vals = {cols[i]: (c.value if c.value is not None else '')
                for i, c in enumerate(r) if i < len(cols) and cols[i]}
        if any(str(v).strip() for k, v in vals.items() if not k.startswith('»')):
            vals['_row'] = r[0].row
            out.append(vals)
    return out

def norm(s):
    """Company names collapse to a comparable key. Punctuation and the entity suffix carry
    no information — 'ABC Foods, L.L.C.' and 'ABC Foods LLC' are the same counterparty, and
    a diff that misses that reports a phantom exit and a phantom entrant from one comma."""
    s = str(s or '').lower()
    s = re.sub(r'[.,&\'"()]', ' ', s)
    s = re.sub(r'\b(llc|l l c|inc|incorporated|corp|corporation|co|company|lp|llp|ltd|'
               r'holdings?|group|enterprises?|partners|restaurants?|foods?|brands?)\b', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

# ------------------------------------------------------------ known names ----
operators = table('OPERATORS')
brands    = table('BRANDS')

op_index = {}
for o in operators:
    keys = [o.get('name', '')] + [a for a in str(o.get('aliases') or '').split('|') if a.strip()]
    for k in keys:
        if norm(k):
            op_index.setdefault(norm(k), o)
brand_index = {norm(b.get('name', '')): b for b in brands if norm(b.get('name', ''))}

def resolve(name, index, cutoff=0.86):
    """Exact on the normalised key, else a conservative fuzzy suggestion. A fuzzy hit is
    reported, never applied — the alias column is where a human confirms it."""
    n = norm(name)
    if not n:
        return None, 'blank', ''
    if n in index:
        return index[n], 'exact', ''
    near = difflib.get_close_matches(n, list(index), n=1, cutoff=cutoff)
    if near:
        return index[near[0]], 'fuzzy', index[near[0]].get('name', '')
    return None, 'unmatched', ''

# ------------------------------------------------------- what CSW already has ----
def load_csw(f):
    p = os.path.join(ROOT, 'data', f)
    return json.load(open(p, encoding='utf-8')) if os.path.exists(p) else None
csw_brands = load_csw('brands.json') or []
csw_metric = {}
for b in csw_brands:
    for k, v in (b.get('metrics') or {}).items():
        csw_metric[(norm(b.get('name', '')), k)] = v

METRIC_MAP = {'auv_revenue': 'auvUsd', 'unit_count': 'usUnits'}

# --------------------------------------------------------------------- sweep ----
inbox = table('INBOX')
roster = table('FDD_ROSTER')
events = table('EVENTS')

report = {'new': [], 'unmatched_operators': [], 'fuzzy': [], 'contradicts': [],
          'ready': [], 'needs_source': [], 'ready_no_source': []}
assign = []            # (tab, row, column-name, value)
next_in = 1 + max([int(re.sub(r'\D', '', str(r.get('» row_id') or '0')) or 0) for r in inbox] or [0])

for r in inbox:
    rid = str(r.get('» row_id') or '').strip()
    if not rid:
        rid = f'IN-{next_in:06d}'; next_in += 1
        assign.append(('INBOX', r['_row'], '» row_id', rid))
        r['» row_id'] = rid          # so the findings payload can key on it
    status = str(r.get('» csw_status') or '').strip()
    if str(r.get('what', '')).startswith('EXAMPLE ROW'):
        continue
    if not status:
        report['new'].append((rid, str(r.get('what', ''))[:90]))
        assign.append(('INBOX', r['_row'], '» csw_status', 'review'))

    name = r.get('entity_name', '')
    if name:
        hit, how, sugg = resolve(name, op_index)
        if how == 'exact':
            assign.append(('INBOX', r['_row'], '» op_id', hit.get('» op_id', '')))
        elif how == 'fuzzy':
            report['fuzzy'].append((rid, name, sugg))
        elif how == 'unmatched':
            bhit, bhow, _ = resolve(name, brand_index)
            if bhow != 'exact':
                report['unmatched_operators'].append((rid, name))

    if not str(r.get('source_url') or '').strip() and not str(r.get('source_name') or '').strip():
        report['needs_source'].append((rid, str(r.get('what', ''))[:70]))

    # contradiction against what the site publishes today
    cat = str(r.get('category') or '')
    mk = METRIC_MAP.get(cat)
    if mk and r.get('value'):
        key = (norm(r.get('brand') or name), mk)
        if key in csw_metric:
            try:
                new = float(re.sub(r'[^\d.\-]', '', str(r['value'])))
                old = float(csw_metric[key])
                if old and abs(new - old) / old > 0.02:
                    report['contradicts'].append((rid, r.get('brand') or name, mk, old, new))
            except ValueError:
                pass

    # Opt-in, never inferred. A row leaves the notebook because you said so, not because it
    # looked confident enough — the whole point of a notebook is that being sloppy in it is
    # safe, and a heuristic that promotes your *confident* sloppiness breaks exactly that.
    if str(r.get('ready') or '').strip().lower() == 'yes' and status not in ('submitted', 'live'):
        if not str(r.get('source_url') or '').strip():
            report['ready_no_source'].append((rid, str(r.get('what', ''))[:70]))
        else:
            report['ready'].append(r)

# ------------------------------------------------------- findings payload ----
items = []
for r in report['ready']:
    cat = str(r.get('category') or '')
    items.append({
        'target_table': 'public.facts' if cat in METRIC_MAP else 'public.entity_notes',
        'title': str(r.get('what', ''))[:120],
        'entity_label': r.get('entity_name', '') or r.get('brand', ''),
        'operation': 'insert',
        'confidence': r.get('confidence', 'medium'),
        'dedupe_key': str(r.get('» row_id') or '').lower(),
        'rationale': f"From the research database, {r.get('date_found','')}, found by "
                     f"{r.get('found_by','')}. {str(r.get('notes','') or '')[:200]}",
        'payload': {'_needs_review': 'metric_key, subject_id and period must be confirmed '
                                     'against the live schema before submitting',
                    'value_numeric': r.get('value', ''), 'period_label': r.get('period', '')},
        'sources': [{'publisher': r.get('source_name', ''),
                     'title': str(r.get('what', ''))[:120],
                     'url': r.get('source_url', ''),
                     'date_label': r.get('period', ''),
                     'quote': str(r.get('quote', '') or '')[:500]}],
    })

# ------------------------------------------------------------------ output ----
W = sys.stdout.write
W(f'\nsweep — {os.path.basename(PATH)} — {TODAY}\n')
W(f'  INBOX {len(inbox)} rows · OPERATORS {len(operators)} · BRANDS {len(brands)} · '
  f'FDD_ROSTER {len(roster)} · EVENTS {len(events)}\n\n')

def block(title, rows, fmt):
    if not rows:
        return
    W(f'  {title} ({len(rows)})\n')
    for x in rows[:15]:
        W('    ' + fmt(x) + '\n')
    if len(rows) > 15:
        W(f'    … and {len(rows)-15} more\n')
    W('\n')

block('new since last sweep', report['new'], lambda x: f'{x[0]}  {x[1]}')
block('operator names I could not match — add an alias, or they are genuinely new',
      report['unmatched_operators'], lambda x: f'{x[0]}  "{x[1]}"')
block('probable matches — confirm by adding to the alias column, I will not guess',
      report['fuzzy'], lambda x: f'{x[0]}  "{x[1]}"  ->  {x[2]}?')
block('CONTRADICTS the live site', report['contradicts'],
      lambda x: f'{x[0]}  {x[1]} {x[2]}: site {x[3]:,} vs database {x[4]:,}')
block('marked ready but has no source — cannot go to CSW without one, so held back',
      report['ready_no_source'], lambda x: f'{x[0]}  {x[1]}')
_named = {x[0] for x in report['ready_no_source']}
block('no source yet (fine — just not submittable)',
      [x for x in report['needs_source'] if x[0] not in _named],
      lambda x: f'{x[0]}  {x[1]}')

W(f'  marked ready and submittable: {len(items)}\n')
if not items:
    W('  (nothing is waiting. Set ready = yes on an INBOX row to promote it.)\n')
if items:
    out = os.path.join(ROOT, 'research', f'sweep-{TODAY}.findings.json')
    json.dump({'batch': {'ref': f'sweep-{TODAY}',
                         'title': f'Research database sweep {TODAY}',
                         'model': 'claude-opus-5',
                         'task_prompt': 'Promote reviewed rows from the research workbook.',
                         'summary': f'{len(items)} rows promoted from INBOX.'},
               'items': items}, open(out, 'w'), indent=2)
    W(f'  wrote {os.path.relpath(out, ROOT)} — review it, then:\n')
    W(f'    node scripts/agent-submit.js {os.path.relpath(out, ROOT)} --dry-run\n')

if WRITE and assign:
    for tab, row, col, val in assign:
        ws = wb[tab]
        cols = [c.value for c in ws[2]]
        if col in cols:
            ws.cell(row=row, column=cols.index(col) + 1, value=val)
    dest = PATH.replace('.xlsx', '-swept.xlsx')
    wb.save(dest)
    W(f'\n  wrote {os.path.relpath(dest, ROOT)} — {len(assign)} machine cells filled, '
      f'your columns untouched\n')
elif assign:
    W(f'\n  {len(assign)} » cells would be filled — re-run with --write to save a copy\n')
W('\n')
