#!/usr/bin/env python3
"""
Creates the CSW Research Database workbook — ONCE.

    pip install openpyxl && python3 scripts/research-db-init.py

This is an `init`, not a build. Once the workbook exists it is the source of truth and
nothing in this repo regenerates it — re-running this script writes to a NEW file rather
than touching one that already has your rows in it. Adding data must never be at risk
from a script run.

The workbook is standalone. It knows about CSW only through two optional machine columns
per tab, and it works perfectly with those left blank forever.

Design rule, and the whole point:

    HUMAN COLUMNS COME FIRST.  MACHINE COLUMNS ARE PREFIXED » AND SIT AT THE FAR RIGHT.

    You never type in a » column. Claude never overwrites a non-» column.

That contract is what lets you and an agent both work in the same file for years without
either one stepping on the other.
"""
import json, os, re, sys, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'research', 'CSW-Research-Database.xlsx')
TODAY = datetime.date.today().isoformat()

def load(name):
    p = os.path.join(ROOT, 'data', name)
    return json.load(open(p, encoding='utf-8')) if os.path.exists(p) else None

brands_j     = load('brands.json') or []
operators_j  = load('operators.json') or []
sources_j    = load('sources.json') or {}
movement_j   = load('movement.json') or {'openings': [], 'closures': []}
transactions_j = load('transactions.json') or {'property': [], 'corporate': []}
news_j       = load('news.json') or []

# ---------------------------------------------------------------- styling ----
INK        = '1A1D21'
ACCENT     = 'B4472A'          # the capture tab — the one you live in
STRUCT     = '1F4E79'          # entity tabs
DOC        = '3F6212'          # documents and extractions
MACHINE_BG = 'E8EAED'
MACHINE_FG = '5F6772'
thin = Side(style='thin', color='D5D8DC')

VOCAB = {
    'entity_type':  'operator,brand,location,person,market,industry,unknown',
    'category':     ('unit_count,auv_revenue,opening,closure,acquisition,divestiture,bankruptcy,'
                     'lawsuit,sale_leaseback,lease,property_sale,development_agreement,refranchising,'
                     'exec_change,financing,distress,expansion_plan,market_entry,market_exit,'
                     'menu_pricing,franchise_terms,other'),
    'confidence':   'confirmed,high,medium,low,rumour',
    'found_by':     'me,claude,broker,news_alert,other',
    'op_type':      ('franchisee,franchisor,private_equity,reit,public_holding,broker,lender,'
                     'developer,landlord,supplier,other'),
    'op_status':    ('active,acquiring,divesting,shrinking,refranchising,newly_formed,stable,'
                     'acquired,restructuring,bankrupt,liquidated,unknown'),
    'yesno':        'yes,no',
    'yesnop':       'yes,no,partial',
    'doc_type':     ('fdd,sec_10k,sec_10q,sec_8k,court_filing,ucc_filing,property_record,'
                     'press_release,news_article,broker_om,research_report,other'),
    'roster_type':  'current,departed',
    'loc_status':   ('open,closed,dark,under_construction,land,for_sale,under_contract,sold,'
                     'converted,demolished,unknown'),
    'priority':     'P0,P1,P2,P3',
    'q_status':     'open,in_progress,blocked,done,dropped',
    'csw_status':   'new,review,queued,submitted,live,rejected,not_for_csw,duplicate',
    'conf_status':  'open,resolved,unresolvable',
    'basis':        ('all_units,franchised_only,company_only,top_quartile,top_half,top_decile,'
                     'mature_units,other'),
    'unit':         'usd,count,pct,sqft,acres,years,ratio,bps,text',
}
COLVOCAB = {
    'entity_type': 'entity_type', 'category': 'category', 'confidence': 'confidence',
    'found_by': 'found_by', 'type': 'op_type', 'status': 'op_status', 'watch': 'yesno',
    'doc_type': 'doc_type', 'have_file': 'yesnop', 'roster_type': 'roster_type',
    'item19_done': 'yesnop', 'item20_done': 'yesnop', 'roster_done': 'yesnop',
    'priority': 'priority', 'unit': 'unit', 'basis': 'basis',
    'is_chicken': 'yesno', 'still_operating': 'yesno',
    '» csw_status': 'csw_status', '» conflict_status': 'conf_status',
}
# q_status and loc_status are applied by explicit tab+column below
SPECIAL_VOCAB = {('QUEUE', 'status'): 'q_status', ('LOCATIONS', 'status'): 'loc_status'}

WIDE = {'what', 'quote', 'notes', 'thesis', 'detail', 'description', 'why', 'title',
        'address', 'aliases', 'brands', 'states', 'key_people', 'task', '» csw_note',
        'resolution', 'value_a', 'value_b', 'raw_quote', 'terms'}
MED  = {'name', 'entity_name', 'source_url', 'url', 'legal_name', 'operator_name',
        'brand', 'source_name', 'city', 'parent_company', 'website', 'role', 'file_location'}

wb = Workbook(); wb.remove(wb.active)
TABS = []

def sheet(name, colour, blurb, cols, rows=None, note=None):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = colour
    n = max(len(cols), 3)

    ws.cell(row=1, column=1, value=blurb)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1)
    c.font = Font(size=10, italic=True, color=colour)
    c.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[1].height = 34

    for i, col in enumerate(cols, start=1):
        mach = col.startswith('»')
        h = ws.cell(row=2, column=i, value=col)
        h.font = Font(bold=not mach, size=10, color=MACHINE_FG if mach else 'FFFFFF',
                      italic=mach)
        h.fill = PatternFill('solid', fgColor=MACHINE_BG if mach else colour)
        h.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
        ws.column_dimensions[get_column_letter(i)].width = (
            52 if col in WIDE else 24 if col in MED else max(11, min(20, len(col) + 3)))
    ws.row_dimensions[2].height = 28

    for ri, row in enumerate(rows or [], start=3):
        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ''))
            cell.font = Font(size=10, color=MACHINE_FG if col.startswith('»') else INK)
            cell.alignment = Alignment(wrap_text=col in WIDE, vertical='top')
            cell.border = Border(bottom=thin)
            if col.startswith('»'):
                cell.fill = PatternFill('solid', fgColor='F6F7F8')

    ws.freeze_panes = 'C3' if len(cols) > 6 else 'A3'
    last = get_column_letter(len(cols))
    ws.auto_filter.ref = f'A2:{last}{max(len(rows or []) + 2, 3)}'

    for i, col in enumerate(cols, start=1):
        key = SPECIAL_VOCAB.get((name, col)) or COLVOCAB.get(col)
        if not key:
            continue
        dv = DataValidation(type='list', formula1='"' + VOCAB[key] + '"',
                            allow_blank=True, showDropDown=False)
        dv.errorTitle, dv.error = 'Pick from the list', 'Or leave it blank — blank is always fine.'
        ws.add_data_validation(dv)
        L = get_column_letter(i)
        dv.add(f'{L}3:{L}5000')
    TABS.append((name, len(cols), len(rows or [])))
    return ws

# =============================================================== START HERE ==
ws = wb.create_sheet('START HERE')
ws.sheet_properties.tabColor = INK
lines = [
    ('CSW RESEARCH DATABASE', 'h1'),
    ('Everything we learn about the chicken restaurant industry, in one file, forever.', 'sub'),
    ('', ''),
    ('THE ONLY RULE YOU NEED', 'h2'),
    ('Found something? Put it in INBOX. One row. Fill in as much as you know and leave the rest '
     'blank — a row with just a sentence and a link is a good row. Do not stop to work out which '
     'tab it belongs in, or look up an ID, or check whether we already have it. That is my job '
     'later, and doing it at capture time is what kills databases like this one.', 'p'),
    ('', ''),
    ('WHO OWNS WHICH COLUMNS', 'h2'),
    ('Columns starting with » are mine. They are grey, they sit at the far right of every tab, '
     'and you never need to touch them. Everything else is yours, and I will never overwrite it.', 'p'),
    ('', ''),
    ('THE TABS, IN THE ORDER YOU WILL ACTUALLY USE THEM', 'h2'),
    ('INBOX          Where 90% of what you add goes. Anything, in any state of certainty.', 'mono'),
    ('OPERATORS      Franchisee groups, PE firms, REITs, landlords. Grows over the years.', 'mono'),
    ('BRANDS         Chicken chains, and the non-chicken brands operators also run.', 'mono'),
    ('LOCATIONS      Individual restaurants. Address-first.', 'mono'),
    ('PEOPLE         Executives and decision makers — who to call, and who moved where.', 'mono'),
    ('EVENTS         Dated things that happened: deals, closures, filings, sale-leasebacks.', 'mono'),
    ('SOURCES        Documents and links. An FDD you hold is a row here.', 'mono'),
    ('FDD_ROSTER     Franchisee lists pulled out of FDDs, one row per franchisee per year.', 'mono'),
    ('FDD_UNITS      Item 20 unit-count tables. Opened, closed, transferred, by state by year.', 'mono'),
    ('CONFLICTS      Two sources that disagree. Never resolve one by deleting a row.', 'mono'),
    ('QUEUE          What to look into next.', 'mono'),
    ('» CSW_LOG      Mine. What has been pushed to the website and what came back.', 'mono'),
    ('', ''),
    ('THE SWEEP', 'h2'),
    ('Whenever you want, tell me to run a sweep. I read this file, work out which rows are new, '
     'match names against operators we already know, flag anything that contradicts what the site '
     'says, and turn the good rows into a submission for the CSW review queue. Nothing goes live '
     'without you approving it in the desk and pressing Publish.', 'p'),
    ('', ''),
    ('So the year of research you do in here is not research I have to redo. That is the point of '
     'the file.', 'p'),
    ('', ''),
    ('ON FDDs — WORTH KNOWING BEFORE YOU START', 'h2'),
    ('A Franchise Disclosure Document contains a complete list of every franchisee, and a second '
     'list of every franchisee who left in the last year, with phone numbers. The FTC requires '
     'both, annually, from every franchisor (16 CFR 436.5(t)).', 'p'),
    ('Collect the same brand across several years and put each list in FDD_ROSTER, and the '
     'comparison answers things nobody else can: who is growing, who is shrinking two years '
     'running, who quietly disappeared, and who bought their units. That is why the roster tab '
     'stores whole lists rather than interesting bits — a name missing from this year\'s list is '
     'only meaningful if last year\'s list was complete.', 'p'),
    ('', ''),
    ('WHAT IS ALREADY IN HERE', 'h2'),
    (f'Seeded on {TODAY} from the live site so you are not starting from an empty file: '
     f'{len(brands_j)} brands, operators, sources, transactions and openings. '
     'Everything seeded is marked » from_csw = yes.', 'p'),
]
r = 1
for text, kind in lines:
    c = ws.cell(row=r, column=1, value=text)
    if kind == 'h1':   c.font = Font(size=20, bold=True, color=ACCENT)
    elif kind == 'sub':c.font = Font(size=12, color='5F6772')
    elif kind == 'h2': c.font = Font(size=11, bold=True, color=STRUCT)
    elif kind == 'mono': c.font = Font(size=10, name='Consolas', color=INK)
    else: c.font = Font(size=11, color=INK)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    if kind == 'p' and text:
        ws.row_dimensions[r].height = 15 * (len(text) // 100 + 1)
    r += 1
ws.column_dimensions['A'].width = 108

# ==================================================================== INBOX ==
sheet('INBOX', ACCENT,
      'PUT EVERYTHING HERE. Fill what you know, leave the rest blank. A sentence and a link is a valid row. '
      'Do not look anything up first — matching, de-duplicating and filing is done later, by me.',
      ['date_found', 'found_by', 'what', 'entity_name', 'brand', 'category', 'value', 'unit',
       'period', 'source_url', 'source_name', 'quote', 'confidence', 'entity_type', 'notes',
       '» row_id', '» op_id', '» csw_status', '» csw_note'],
      [{'date_found': TODAY, 'found_by': 'claude',
        'what': 'EXAMPLE ROW — delete me. Yum does not publish a KFC U.S. same-store sales figure; the 2% on the site is the KFC Division number, and that division is ~90% non-U.S. by units.',
        'entity_name': 'KFC', 'brand': 'KFC', 'category': 'other', 'value': '', 'unit': 'text',
        'period': 'YE2025', 'source_url': 'https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=YUM',
        'source_name': 'Yum! Brands Q2 2026 10-Q, FY2025 10-K, Q2 2026 release',
        'quote': 'Checked all three; none discloses a KFC U.S. comparable-sales figure.',
        'confidence': 'confirmed', 'entity_type': 'brand',
        'notes': 'This is what a good row looks like: a claim, who it is about, where it came from, and how sure I am.',
        '» row_id': 'IN-000001', '» csw_status': 'review'}])

# ================================================================ OPERATORS ==
op_rows, seen = [], {}
def add_op(name, **kw):
    k = name.strip().lower()
    if k in seen: return seen[k]
    row = {'name': name, 'aliases': '', 'legal_name': '', 'type': 'franchisee', 'status': '',
           'hq_city': '', 'hq_state': '', 'total_units': '', 'chicken_units': '', 'brands': '',
           'states': '', 'website': '', 'key_people': '', 'parent_company': '', 'watch': '',
           'thesis': '', 'notes': '', '» op_id': f'OP-{len(op_rows)+1:04d}',
           '» from_csw': 'yes', '» csw_slug': '', '» last_swept': ''}
    row.update(kw); seen[k] = row; op_rows.append(row); return row

for o in operators_j:
    add_op(o.get('name', ''), legal_name='', type='franchisee',
           status=(o.get('status') or '').lower().replace(' ', '_'),
           hq_state=o.get('hq', ''), total_units=o.get('totalUnits', ''),
           chicken_units=o.get('chickenUnits', ''),
           brands=', '.join(o.get('brands', [])), states=', '.join(o.get('geography', [])),
           thesis=(o.get('analysis', '') or '')[:600],
           **{'» csw_slug': o.get('slug', '')})
for t in transactions_j.get('corporate', []):
    if t.get('acquirer'):
        add_op(t['acquirer'], type='private_equity' if 'capital' in t['acquirer'].lower() or 'partners' in t['acquirer'].lower() else 'other',
               notes=f"Appears in CSW transactions as an acquirer ({t.get('date','')}).",
               **{'» from_csw': 'yes', '» csw_slug': ''})

sheet('OPERATORS', STRUCT,
      'One row per company. Type the name the way you found it; put every other spelling in `aliases`, '
      'separated by | — that is what lets a 2022 FDD match a 2026 one. `watch` = your watchlist. `thesis` = your view.',
      ['name', 'aliases', 'legal_name', 'type', 'status', 'hq_city', 'hq_state', 'total_units',
       'chicken_units', 'brands', 'states', 'website', 'key_people', 'parent_company',
       'watch', 'thesis', 'notes', '» op_id', '» from_csw', '» csw_slug', '» last_swept'],
      op_rows)

# =================================================================== BRANDS ==
br_rows = []
for b in brands_j:
    st = b.get('stats', {}) or {}
    br_rows.append({
        'name': b.get('name', ''), 'aliases': '', 'is_chicken': 'yes',
        'franchisor_legal_name': '', 'parent_company': b.get('parent', ''),
        'us_units': (st.get('usUnits') or {}).get('v', ''),
        'auv': (b.get('metrics', {}) or {}).get('auvUsd', ''),
        'fdd_years_held': '', 'fiscal_year_end': '', 'segment': '', 'notes': '',
        '» brand_id': f"BR-{b.get('slug','')}", '» from_csw': 'yes',
        '» csw_slug': b.get('slug', ''), '» last_swept': ''})
sheet('BRANDS', STRUCT,
      'Chicken chains, plus the non-chicken brands your operators also run — a franchisee\'s chicken exposure '
      'only means something against the rest of their portfolio.',
      ['name', 'aliases', 'is_chicken', 'franchisor_legal_name', 'parent_company', 'us_units',
       'auv', 'fdd_years_held', 'fiscal_year_end', 'segment', 'notes',
       '» brand_id', '» from_csw', '» csw_slug', '» last_swept'], br_rows)

# ================================================================ LOCATIONS ==
loc_rows = []
def add_loc(**kw):
    row = {'address': '', 'city': '', 'state': '', 'zip': '', 'brand': '', 'operator_name': '',
           'status': '', 'opened': '', 'closed': '', 'sqft': '', 'drive_thru': '',
           'prior_use': '', 'owned_or_leased': '', 'landlord': '', 'last_sale_price': '',
           'last_sale_date': '', 'cap_rate': '', 'notes': '',
           '» loc_id': f'LOC-{len(loc_rows)+1:05d}', '» op_id': '', '» from_csw': 'yes',
           '» csw_status': ''}
    row.update(kw); loc_rows.append(row)
for o in movement_j.get('openings', []):
    add_loc(address=o.get('location', ''), brand=o.get('brandName', ''), status='open',
            opened=o.get('date', ''), notes=o.get('detail', ''))
for t in transactions_j.get('property', []):
    add_loc(address=t.get('location', ''), brand=t.get('brand', ''), status='open',
            last_sale_price=t.get('price', ''), last_sale_date=t.get('date', ''),
            cap_rate=t.get('capRate', ''), notes=t.get('detail', '')[:400])
sheet('LOCATIONS', STRUCT,
      'One row per physical restaurant. The row survives a change of brand or operator — a former KFC that '
      'reopens as a Dave\'s is the same box, and second-generation supply is the site\'s whole thesis.',
      ['address', 'city', 'state', 'zip', 'brand', 'operator_name', 'status', 'opened', 'closed',
       'sqft', 'drive_thru', 'prior_use', 'owned_or_leased', 'landlord', 'last_sale_price',
       'last_sale_date', 'cap_rate', 'notes',
       '» loc_id', '» op_id', '» from_csw', '» csw_status'], loc_rows)

# =================================================================== PEOPLE ==
sheet('PEOPLE', STRUCT,
      'Executives and decision makers. Who signs a deal, who just moved, who to call. A departure is often '
      'the earliest signal an operator is in trouble or about to sell.',
      ['name', 'role', 'company', 'brand', 'email', 'phone', 'linkedin', 'start_date',
       'end_date', 'previous_company', 'still_operating', 'source_url', 'notes',
       '» person_id', '» op_id', '» last_swept'], [])

# =================================================================== EVENTS ==
ev_rows = []
for t in transactions_j.get('corporate', []):
    ev_rows.append({'date': t.get('date', ''), 'event_type': 'acquisition',
                    'headline': f"{t.get('acquirer','')} — {t.get('target','')}",
                    'operator_name': t.get('acquirer', ''), 'brand': t.get('target', ''),
                    'counterparty': t.get('target', ''), 'value_usd': t.get('value', ''),
                    'units_involved': '', 'states': '',
                    'detail': t.get('detail', ''), 'source_url': '',
                    'source_name': t.get('src', ''), 'confidence': 'high',
                    '» event_id': f'EV-{len(ev_rows)+1:04d}', '» op_id': '',
                    '» from_csw': 'yes', '» csw_status': 'live'})
for c in movement_j.get('closures', []):
    ev_rows.append({'date': c.get('date', ''), 'event_type': 'closure',
                    'headline': f"{c.get('brandName','')} — {c.get('count','')} closures",
                    'operator_name': '', 'brand': c.get('brandName', ''), 'counterparty': '',
                    'value_usd': '', 'units_involved': c.get('count', ''),
                    'states': c.get('location', ''), 'detail': c.get('detail', ''),
                    'source_url': '', 'source_name': c.get('src', ''), 'confidence': 'high',
                    '» event_id': f'EV-{len(ev_rows)+1:04d}', '» op_id': '',
                    '» from_csw': 'yes', '» csw_status': 'live'})
sheet('EVENTS', STRUCT,
      'Dated things that happened: acquisitions, closures, bankruptcies, sale-leasebacks, development '
      'agreements, refranchising. One row per event. Promote these out of INBOX once they are firm.',
      ['date', 'event_type', 'headline', 'operator_name', 'brand', 'counterparty', 'value_usd',
       'units_involved', 'states', 'detail', 'source_url', 'source_name', 'confidence',
       '» event_id', '» op_id', '» from_csw', '» csw_status'], ev_rows)

# ================================================================== SOURCES ==
src_rows = []
for key, s in sources_j.items():
    src_rows.append({'title': s.get('title', ''), 'publisher': s.get('pub', ''),
                     'url': s.get('url', ''), 'date': s.get('date', ''),
                     'doc_type': 'news_article', 'brand': '', 'have_file': 'no',
                     'file_location': '', 'notes': '',
                     '» source_id': f'SRC-{key}', '» from_csw': 'yes'})
KNOWN_FDDS = [
    ("Zaxby's",             "Zaxby's SPE Franchisor LLC",              '2026-04-24', 'WI file 641240'),
    ('Bojangles',           'Bojangles Opco, LLC',                     '2026-04-20', ''),
    ('Chicken Salad Chick', 'Simply Southern Restaurant Group, LLC',   '2026-04-21', ''),
    ('Slim Chickens',       "Slim Chicken's Development Company, LLC", '2026-04-29', ''),
    ("Dave's Hot Chicken",  "Dave's Hot Chicken Franchise Co. SPV LLC",'2026-05-04', ''),
]
for name, registrant, eff, extra in KNOWN_FDDS:
    src_rows.append({'title': f'{name} FDD, effective {eff}', 'publisher': registrant,
                     'url': 'https://apps.dfi.wi.gov/apps/FranchiseSearch/MainSearch.aspx',
                     'date': eff, 'doc_type': 'fdd', 'brand': name, 'have_file': 'no',
                     'file_location': '',
                     'notes': f'Registration confirmed 2026-09-01. {extra} Document not obtained — '
                              'this is why the AUV on the site is stale.',
                     '» source_id': f'DOC-{re.sub(r"[^a-z]+","-",name.lower()).strip("-")}-{eff[:4]}',
                     '» from_csw': 'no'})
sheet('SOURCES', DOC,
      'Every document and link. An FDD you have downloaded is a row here with have_file = yes and the '
      'file path in file_location. Rows with doc_type = fdd and have_file = no are the shopping list.',
      ['title', 'publisher', 'url', 'date', 'doc_type', 'brand', 'have_file', 'file_location',
       'item19_done', 'item20_done', 'roster_done', 'notes', '» source_id', '» from_csw'],
      src_rows)

# =============================================================== FDD_ROSTER ==
sheet('FDD_ROSTER', DOC,
      'Franchisee lists out of FDDs — one row per franchisee per document. Paste WHOLE lists, including the '
      'boring ones: a name missing from this year is only meaningful if last year was complete. '
      'roster_type = departed is the FTC-required list of franchisees who left in the last fiscal year, '
      'with phone numbers. That list is the single most valuable page in the document.',
      ['brand', 'fdd_year', 'fdd_date', 'roster_type', 'operator_name', 'address', 'city',
       'state', 'zip', 'phone', 'units', 'departure_reason', 'page', 'notes',
       '» roster_id', '» source_id', '» op_id', '» loc_id', '» match_note'], [])

# ================================================================ FDD_UNITS ==
sheet('FDD_UNITS', DOC,
      'Item 20 unit-count tables — one row per state per fiscal year. Columns match the FDD\'s own headings so '
      'you can transcribe straight across. Terminated / non-renewed / reacquired and transfers are where '
      'distress and consolidation show up as numbers rather than as a story.',
      ['brand', 'fdd_year', 'fiscal_year', 'table_no', 'state', 'outlet_type', 'at_start',
       'opened', 'terminated', 'non_renewed', 'reacquired', 'ceased_other', 'at_end',
       'transfers', 'projected_openings', 'page', 'notes', '» source_id'], [])

# ================================================================ CONFLICTS ==
sheet('CONFLICTS', DOC,
      'Two sources that disagree. Keep both. Resolving one by deleting a row is how a database quietly '
      'starts lying — and the spread itself is often the finding.',
      ['subject', 'metric', 'value_a', 'source_a', 'value_b', 'source_b', 'spread', 'why_it_matters',
       'resolution', 'notes', '» conflict_id', '» conflict_status'],
      [{'subject': "Zaxby's", 'metric': 'AUV', 'value_a': '$2,847,345', 'source_a': 'FDD aggregator site A',
        'value_b': '$2,544,354', 'source_b': 'FDD aggregator site B', 'spread': '$303,000',
        'why_it_matters': 'This figure decides a rank on the site. Neither source is primary.',
        'resolution': 'Only Item 19 of the 2026-04-24 FDD settles it. Reachable is not sourced.',
        'notes': '', '» conflict_id': 'CF-0001', '» conflict_status': 'open'}])

# ==================================================================== QUEUE ==
q = [
 ('P0', 'Download the five 2026 FDDs (Zaxby\'s, Bojangles, Chicken Salad Chick, Slim Chickens, Dave\'s Hot Chicken)',
  'Eight of the eleven overdue figures on the live site are FDD Item 19 AUVs. One download session fixes all of them.',
  'me', 'cards.commerce.state.mn.us — free, no account. Search the registrant legal name in SOURCES, not the brand name.'),
 ('P0', 'Decide what to do about the KFC U.S. same-store sales input',
  'A scoring input measures something other than what its label says. No freshness check can catch that.',
  'me', 'Claude will compute the score both ways first. 10 minutes.'),
 ('P1', 'Download Popeyes FDDs 2021-2026',
  'Proves the year-over-year roster comparison end to end, on the brand where we already track five operators.',
  'me', 'Same source. Item 20 lists are the target here, not the AUV.'),
 ('P1', 'Refresh El Pollo Loco Q4 comps', 'Overdue on the live site and self-serviceable.',
  'claude', 'EPL files with the SEC; data.sec.gov is reachable. No help needed.'),
 ('P2', 'Work out whether California DFPI DOCQNET can serve FDDs programmatically',
  'If it can, the download step disappears and the FDD programme becomes self-serve.',
  'claude', 'docqnet.dfpi.ca.gov responds. Unproven that it exposes full documents. Timeboxed.'),
]
sheet('QUEUE', DOC,
      'What to look into next, ordered by value rather than by ease. Add a row whenever you think of something.',
      ['priority', 'task', 'why', 'owner', 'how', 'status', 'opened', 'closed', 'notes', '» task_id'],
      [{'priority': p, 'task': t, 'why': w, 'owner': o, 'how': h, 'status': 'open',
        'opened': TODAY, 'closed': '', 'notes': '', '» task_id': f'Q-{i+1:03d}'}
       for i, (p, t, w, o, h) in enumerate(q)])

# ================================================================= CSW_LOG ==
sheet('» CSW_LOG', '5F6772',
      'Mine. A line per sweep: what was submitted to the CSW review queue, when, and what came back. '
      'You never need to edit this — it is here so the round trip is auditable.',
      ['swept_on', 'batch_ref', 'source_tab', 'source_row_id', 'subject', 'proposed_change',
       'target_table', 'outcome', 'reviewed_by', 'live_on_site', 'notes'], [])

os.makedirs(os.path.dirname(OUT), exist_ok=True)
if os.path.exists(OUT) and '--force' not in sys.argv:
    sys.exit(f'refusing to overwrite {OUT} — it may have your rows in it. Pass --force if you are sure.')
wb.save(OUT)
print('wrote', os.path.relpath(OUT, ROOT))
for n, c, r in TABS:
    print(f'  {n:<14} {c:>3} cols  {r:>4} seeded rows')
